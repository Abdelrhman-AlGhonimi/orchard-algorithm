import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import os
from datetime import datetime
from openpyxl import load_workbook

# Get current directory and output folder
script_dir = os.path.dirname(os.path.abspath(__file__))
output_dir = script_dir  # Save in the same directory as the script

# Define file paths
file_paths = {
    'greedy': r'd:\\abdo1\\VS Code\\test\\Exels\\greedy_scheduling_schedule.xlsx',
    'random': r'd:\\abdo1\\VS Code\\test\\Exels\\random_assignment_schedule.xlsx',
    'annealing': r'd:\\abdo1\\VS Code\\test\\Exels\\simulated_annealing_schedule.xlsx',
    'orchard': r'd:\\abdo1\\VS Code\\test\\Exels\\Orchard_algorithm.xlsx',
    'genetic': r'd:\\abdo1\\VS Code\\test\\Exels\\genetic_scheduling_schedule.xlsx'
}

# Define colors for each algorithm
colors = {
    'greedy': 'blue',
    'random': 'red',
    'annealing': 'green',
    'orchard': 'purple',
    'genetic': 'orange'
}

# Define algorithm names for legend
algo_names = {
    'greedy': 'Greedy Scheduling',
    'random': 'Random Assignment',
    'annealing': 'Simulated Annealing',
    'orchard': 'Orchard Algorithm',
    'genetic': 'Genetic Algorithm'
}

# Read all Excel files and process data similar to statistics.py
def process_excel_files(file_paths):
    processed_data = {}
    
    for algo_name, file_path in file_paths.items():
        try:
            # Load workbook
            wb = load_workbook(filename=file_path)
            ws = wb.active
            
            # Convert Excel data to list of dictionaries
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:  # Skip empty rows
                    # Try to adapt to different column formats
                    entry = {}
                    
                    # Room - usually the first column
                    if isinstance(row[0], (int, float)) or (isinstance(row[0], str) and row[0].isdigit()):
                        entry['room'] = int(float(row[0]))
                    else:
                        entry['room'] = 1  # Default if not found
                    
                    # Duration - try to find it
                    duration_idx = None
                    for i, val in enumerate(row):
                        if val is not None and isinstance(val, str) and 'min' in val.lower():
                            try:
                                entry['duration'] = int(val.lower().replace('min', '').strip())
                                duration_idx = i
                                break
                            except:
                                pass
                    
                    # If duration not found, look for numeric values
                    if 'duration' not in entry:
                        for i, val in enumerate(row):
                            if isinstance(val, (int, float)) and 10 <= val <= 300:  # Reasonable operation duration in minutes
                                entry['duration'] = int(val)
                                duration_idx = i
                                break
                    
                    # If still not found, use a default
                    if 'duration' not in entry:
                        entry['duration'] = 60  # Default duration
                    
                    # Emergency status - look for boolean or 'Yes'/'No' or 'True'/'False'
                    entry['emergency'] = False  # Default value
                    for i, val in enumerate(row):
                        if val is not None:
                            if isinstance(val, bool):
                                entry['emergency'] = val
                                break
                            elif isinstance(val, str) and val.lower() in ('yes', 'true', 'emergency'):
                                entry['emergency'] = True
                                break
                    
                    # Start time - try to find time format
                    time_found = False
                    for i, val in enumerate(row):
                        if val is not None and isinstance(val, str):
                            # Check for time formats like '8:00 AM', '08:00'
                            if ':' in val:
                                entry['start_time'] = val
                                time_found = True
                                break
                    
                    # If time not found, use position if available in statistics.py
                    if not time_found:
                        for i, val in enumerate(row):
                            if val is not None and i == 5:  # Index 5 is start_time in statistics.py
                                entry['start_time'] = val
                                break
                    
                    # If still no time, use a default start time
                    if 'start_time' not in entry:
                        entry['start_time'] = '8:00 AM'
                    
                    # Add to data list
                    data.append(entry)
            
            # Convert to DataFrame
            df = pd.DataFrame(data)
            
            # Convert start_time to minutes since start of day (as in statistics.py)
            def time_to_minutes(time_str):
                if not isinstance(time_str, str):
                    return 480  # Default to 8:00 AM (480 minutes)
                
                try:
                    # Try different time formats
                    if 'AM' in time_str or 'PM' in time_str:
                        parts = time_str.split()
                        time_part = parts[0].split(':')
                        hour = int(time_part[0])
                        minute = int(time_part[1])
                        period = parts[1]
                        
                        # Convert to 24-hour format
                        if period == 'PM' and hour != 12:
                            hour += 12
                        elif period == 'AM' and hour == 12:
                            hour = 0
                    else:
                        # Try 24-hour format
                        time_part = time_str.split(':')
                        hour = int(time_part[0])
                        minute = int(time_part[1])
                    
                    return hour * 60 + minute
                except:
                    # If parsing fails, default to 8:00 AM
                    return 480
            
            df['start_minutes'] = df.apply(lambda row: time_to_minutes(row['start_time']), axis=1)
            
            processed_data[algo_name] = df
            print(f"Successfully loaded and processed {algo_name} data")
        except Exception as e:
            print(f"Error loading {algo_name} data: {e}")
    
    return processed_data

# 1. Distribution of the number of operations by time
def plot_operations_by_time(datasets):
    plt.figure(figsize=(14, 8))
    
    # Create time bins (every hour from 8:00 AM to 5:00 PM)
    time_bins = np.arange(480, 1020 + 1, 60)  # 8:00 AM to 5:00 PM
    time_labels = []
    for h in range(8, 17):
        if h < 12:
            time_labels.append(f"{h}:00 AM")
        elif h == 12:
            time_labels.append(f"{h}:00 PM")
        else:
            time_labels.append(f"{h-12}:00 PM")
    
    # Group histogram data by algorithm
    for algo_name, df in datasets.items():
        if 'start_minutes' in df.columns and not df.empty:
            # Create histogram with algorithm-specific color
            hist_data, _ = np.histogram(df['start_minutes'], bins=time_bins)
            bin_centers = (time_bins[:-1] + time_bins[1:]) / 2
            plt.plot(bin_centers, hist_data, marker='o', linestyle='-', 
                    color=colors[algo_name], linewidth=2, markersize=8,
                    label=algo_names[algo_name])
    
    # Adjust time ticks to match statistics.py
    plt.xticks(time_bins[:-1], time_labels, rotation=45)
    plt.title('Distribution of Operations by Time', fontsize=16)
    plt.xlabel('Time of Day', fontsize=14)
    plt.ylabel('Number of Operations', fontsize=14)
    plt.grid(True, linestyle='--', alpha=0.7)
    plt.legend(fontsize=12)
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'operations_distribution_comparison.png'))
    plt.close()

# 2. Daily demand for operating rooms
def plot_room_demand(datasets):
    plt.figure(figsize=(14, 8))
    
    # Determine all unique room numbers across all datasets
    all_rooms = set()
    for df in datasets.values():
        if 'room' in df.columns and not df.empty:
            all_rooms.update(df['room'].unique())
    
    all_rooms = sorted(list(all_rooms))
    
    # Set up bar positions
    bar_width = 0.8 / len(datasets)  # Width of each bar
    offsets = np.arange(-(len(datasets)-1)/2, (len(datasets)+1)/2) * bar_width  # Bar offsets
    
    # Plot bars for each algorithm
    for (algo_name, df), offset in zip(datasets.items(), offsets):
        if 'room' in df.columns and 'duration' in df.columns and not df.empty:
            # Calculate total operation time per room
            room_times = df.groupby('room')['duration'].sum()
            
            # Make sure all rooms are represented
            room_data = [room_times.get(room, 0) for room in all_rooms]
            
            # Calculate bar positions
            x_positions = np.arange(len(all_rooms)) + offset
            
            plt.bar(x_positions, room_data, width=bar_width, 
                    color=colors[algo_name], label=algo_names[algo_name])
    
    plt.title('Daily Demand for Operating Rooms', fontsize=16)
    plt.xlabel('Operating Room Number', fontsize=14)
    plt.ylabel('Total Operation Time (minutes)', fontsize=14)
    plt.xticks(np.arange(len(all_rooms)), all_rooms)
    plt.legend(fontsize=12)
    plt.grid(True, axis='y', linestyle='--', alpha=0.7)
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'room_demand_comparison.png'))
    plt.close()

# 3. Analysis of emergency operations by time
def plot_emergency_operations(datasets):
    plt.figure(figsize=(14, 8))
    
    # Create time bins (every hour from 8:00 AM to 5:00 PM)
    time_bins = np.arange(480, 1020 + 1, 60)  # 8:00 AM to 5:00 PM
    time_labels = []
    for h in range(8, 17):
        if h < 12:
            time_labels.append(f"{h}:00 AM")
        elif h == 12:
            time_labels.append(f"{h}:00 PM")
        else:
            time_labels.append(f"{h-12}:00 PM")
    
    # Plot emergency operations for each algorithm
    for algo_name, df in datasets.items():
        if 'emergency' in df.columns and 'start_minutes' in df.columns and not df.empty:
            # Filter emergency operations
            emergency_ops = df[df['emergency'] == True]
            
            if not emergency_ops.empty:
                # Create histogram data
                hist_data, _ = np.histogram(emergency_ops['start_minutes'], bins=time_bins)
                bin_centers = (time_bins[:-1] + time_bins[1:]) / 2
                
                plt.plot(bin_centers, hist_data, marker='s', linestyle='-', 
                        color=colors[algo_name], linewidth=2, markersize=8,
                        label=algo_names[algo_name])
    
    # Adjust time ticks to match statistics.py
    plt.xticks(time_bins[:-1], time_labels, rotation=45)
    plt.title('Emergency Operations by Time', fontsize=16)
    plt.xlabel('Time of Day', fontsize=14)
    plt.ylabel('Number of Emergency Operations', fontsize=14)
    plt.grid(True, linestyle='--', alpha=0.7)
    plt.legend(fontsize=12)
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'emergency_distribution_comparison.png'))
    plt.close()

# Main function
def main():
    print("Starting algorithm comparison...")
    
    # Process the Excel files
    datasets = process_excel_files(file_paths)
    
    # Generate plots
    print("\nGenerating distribution of operations by time...")
    plot_operations_by_time(datasets)
    
    print("Generating daily demand for operating rooms...")
    plot_room_demand(datasets)
    
    print("Generating analysis of emergency operations by time...")
    plot_emergency_operations(datasets)
    
    print(f"\nAll comparison plots have been generated and saved in: {output_dir}")
    print("1. operations_distribution_comparison.png - Distribution of operations by time")
    print("2. room_demand_comparison.png - Daily demand for operating rooms")
    print("3. emergency_distribution_comparison.png - Emergency operations by time")

if __name__ == "__main__":
    main()